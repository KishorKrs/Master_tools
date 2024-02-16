import os
import sys
from pandas import read_csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from csv import reader, writer, QUOTE_ALL, QUOTE_MINIMAL

sys.path.append("../")
from helper.main import Main


class CSV(Main):
    def prepare_init(self) -> None:
        pass

    def read(self, fpath, encoding="utf-8"):
        if not self.futil.is_file(fpath):
            msg = "File doesnot Exists. Terminating"
            self.log.error(msg)
            raise Exception(msg)

        output_data = None
        with open(fpath, "r", encoding=encoding) as input_file:
            output_data = input_file.readlines()
        return output_data

    def write_colored(self, fpath, data, header):
        # Create a workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active

        # Write data to worksheet
        for i, row in enumerate(data, start=0):
            if i % 2 == 0:
                idx = self.compare_row(data[i], data[i + 1])
            for j, val in enumerate(row, start=0):
                cell = ws.cell(row=i + 1, column=j + 1)
                cell.value = val

                if j in idx:
                    cell.font = Font(color="00FF0000")  # Red color
                    cell.fill = PatternFill("solid", start_color="FFFF00")

        # Save the workbook
        wb.save(fpath)

    def compare_row(self, row1, row2):
        idx = []
        if len(row1) >= len(row2):
            data_row1 = row1
            data_row2 = row2
        else:
            data_row1 = row2
            data_row2 = row1

        for i, val in enumerate(data_row1):
            try:
                if data_row2[i] != val:
                    idx.append(i)
            except:
                idx.append(i)

        return idx

    def write(self, fpath, data_list, encoding="utf-8", quote_all=True):
        if not data_list:
            self.log.warning("No Data Present in the data_list. Returning")
            return

        dir_name = os.path.dirname(fpath)
        if not self.futil.is_dir(dir_name):
            self.futil.make_dir(dir_name)
            self.log.info("Created Parent Dir to store Output File")

        with open(fpath, "w", encoding=encoding, newline="") as output_file:
            csv_writer = writer(output_file, QUOTE_ALL if quote_all else QUOTE_MINIMAL)
            for d in data_list:
                csv_writer.writerow(d)

        self.log.info("Process of writing to CSV File completed")

    def is_same(self, fpath1, fpath2, encoding="utf-8"):
        df1 = read_csv(fpath1, encoding=encoding)
        df2 = read_csv(fpath2, encoding=encoding)
        df_diff = df1.compare(df2)
        all_matched = df_diff.empty
        return all_matched

    def compare_files(self, fpath1, fpath2, diff_by=0, encoding="utf-8", skip_header=True, output_diff_path=None):
        """Compares two CSV Files

        Args
            fpath1: File path for 1st CSV File
            fpath2: File path for 2nd CSV File
            diff_by: column_id. Think it as ID of unique element. Default is 0 which means whole will be compared
            output_diff: Flag to tell if difference should be output to file or not
        """
        if fpath1 == fpath2:
            self.log.info("Given files are Same File.")
            return True

        container = {}
        is_different = False
        header = None

        with open(fpath1, "r", encoding=encoding) as f1:
            f1_reader = reader(f1)
            for idx, row in enumerate(f1_reader):
                if not idx and skip_header:
                    header = row
                    continue
                unique_field = row[diff_by]
                if unique_field not in container:
                    container[unique_field] = {"data1": row, "matched": False, "count1": 1, "contains_data": ["data1"]}
                else:
                    container[unique_field]["count1"] += 1

        with open(fpath2, "r", encoding=encoding) as f2:
            f2_reader = reader(f2)
            for idx, row in enumerate(f2_reader):
                if not idx and skip_header:
                    continue

                unique_field = row[diff_by]
                if unique_field in container:
                    if container[unique_field]["data1"] == row:
                        del container[unique_field]
                    elif "data2" in container[unique_field]:
                        container[unique_field]["count2"] += 1
                    else:
                        container[unique_field]["data2"] = row
                        container[unique_field]["contains_data"].append("data2")
                        is_different = True
                else:
                    container[unique_field] = {"data2": row, "matched": False, "count2": 1, "contains_data": ["data2"]}
                    is_different = True

        if not output_diff_path:
            return is_different

        if not output_diff_path.endswith(".xlsx"):
            output_diff_path = self.futil.change_extension(output_diff_path, "xlsx")

        data_list = []
        wb = Workbook()
        ws = wb.active
        row_idx = 1

        green = "90eda0"
        blue = "91b3fd"

        for key in container:
            if key == "\x1a":
                continue

            data_containes = container[key]["contains_data"]
            if len(data_containes) == 1:
                has_data = data_containes[0]
                color = green if has_data == "data1" else blue
                row = container[key][has_data]
                row.insert(0, "file1" if has_data == "data1" else "file2")
                for j, val in enumerate(row, start=0):
                    cell = ws.cell(row=row_idx, column=j + 1)
                    cell.value = val
                    cell.fill = PatternFill("solid", start_color=color)
                row_idx += 1

            else:
                row1 = container[key]["data1"]
                row2 = container[key]["data2"]
                row1.insert(0, "file1")
                row2.insert(0, "file2")

                # Write data to worksheet
                idx = self.compare_row(row1, row2)
                for i, row in enumerate([row1, row2]):
                    for j, val in enumerate(row, start=0):
                        cell = ws.cell(row=row_idx, column=j + 1)
                        cell.value = val

                        if j in idx:
                            cell.font = Font(color="00FF0000")  # Red color
                            cell.fill = PatternFill("solid", start_color="FFFF00")
                    row_idx += 1

        # Save the workbook
        wb.save(output_diff_path)

        return is_different

    def split_csv(self, fpath, col_idx=0, skip_header=True, encoding="utf-8"):
        if not self.futil.is_file(fpath):
            self.log.warning("Given file doesnot exists!")
            return

        fdir, fname = self.futil.split_names(fpath)
        fname = fname.split(".")[0]
        output_dir = fpath.split(".csv")[0]
        self.futil.make_dir(output_dir)
        part_fpath = self.futil.path_join(output_dir, fname)

        data_container = {}
        with open(fpath, "r", encoding=encoding) as input_file:
            input_reader = reader(input_file)
            for idx, row in enumerate(input_reader):
                if not idx and skip_header:
                    continue

                if len(row) <= col_idx:
                    continue

                identifier = row[col_idx]
                if identifier not in data_container:
                    data_container[identifier] = []
                data_container[identifier].append(row)

        for key in data_container:
            splited_fpath = f"{part_fpath}_{key}.csv"
            with open(splited_fpath, "w", encoding=encoding, newline="") as output_file:
                input_writer = writer(output_file, QUOTE_ALL)
                for row in data_container[key]:
                    input_writer.writerow(row)
